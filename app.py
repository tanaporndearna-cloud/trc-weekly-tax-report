"""
Weekly Tax Report — TRC Motorsport
Streamlit App  |  github.com/tanaporndearna-cloud/trc-weekly-tax-report
"""

import base64
import io
import json
import re
from datetime import date, timedelta

import anthropic
import gspread
import pandas as pd
import requests
import streamlit as st
from google.auth.transport.requests import Request as GoogleAuthRequest
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────
SHEET_ID   = "1lD6YrCoSbA5RI79PvtG2WtWTVd0q83xvPmESLDMcu-0"
FORM_SHEET = "การตอบแบบฟอร์ม 1"

BRANCH_MAP = {
    "BNA": 3,  "BNB": 4,  "BNC": 7,  "BND": 5,  "CJH": 10,
    "PAN": 9,  "PRS": 6,  "PTA": 1,  "PTB": 2,  "PTC": 8,
    "PTY": 11, "TBY": 16, "TCB": 13, "TEM": 12, "TLX": 17,
    "TNW": 14, "TNY": 15, "TSP": 18,
}

PRODUCT_ABBR = {
    "JAZZ": "JZ", "VIOS": "VO", "CIVIC": "CV", "ATTRAGE": "AT",
    "SWIFT": "SW", "CIAZ": "CZ", "HR-V": "HR", "YARIS ATIV": "YA",
    "YARIS": "YA",
}

DAY_TH = ["จันทร์", "อังคาร", "พุธ", "พฤหัส", "ศุกร์", "เสาร์", "อาทิตย์"]

DISPLAY_COLS = [
    "ว.ด.ป.", "วัน", "เลขที่", "สาขา", "เลขที่เอกสาร",
    "ชื่อลูกค้า", "ที่อยู่", "TAX ID",
    "รายการสินค้า", "จำนวน", "ราคา/หน่วย",
    "รวม", "VAT 7%", "รวมทั้งสิ้น",
    "ช่องทาง", "อีเมล/จัดส่ง",
]

# ─────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TRC Tax Report",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
    div[data-testid="metric-container"] {
        background: white;
        border: 1px solid #e0e8f0;
        border-radius: 10px;
        padding: 12px 18px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def to_thai_date(d: date) -> str:
    return f"{d.day:02d}/{d.month:02d}/{d.year + 543}"


def parse_date_val(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, date):
        return val
    s = str(val).strip().split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return pd.to_datetime(s, format=fmt).date()
        except Exception:
            pass
    return None


def working_day_index(d: date) -> int:
    """Count Mon–Sat days from the 1st of the month up to d (inclusive)."""
    return sum(
        1
        for i in range(1, d.day + 1)
        if date(d.year, d.month, i).weekday() != 6
    )


def calc_tax_no(d: date, slot: int) -> int:
    wd   = working_day_index(d)
    yr   = d.year + 543
    base = int(f"{yr}{d.month:02d}000")
    return base + (wd - 1) * 15 + 1 + slot


def get_branch(doc_no: str) -> int | str:
    m = re.match(r"^ABB([A-Z]{2,3})\d", str(doc_no or ""))
    return BRANCH_MAP.get(m.group(1), "") if m else ""


def safe(row, idx, default="") -> str:
    return row[idx].strip() if len(row) > idx and row[idx] else default


def get_drive_file_id(url: str) -> str | None:
    if not url:
        return None
    # Format: https://drive.google.com/file/d/FILE_ID/view
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    # Format: https://drive.google.com/open?id=FILE_ID
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    return None


# ─────────────────────────────────────────────────────────────────
#  GOOGLE AUTH
# ─────────────────────────────────────────────────────────────────

@st.cache_resource
def _get_credentials():
    """Returns (gspread_client, raw_credentials, error_string)"""
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    try:
        info = dict(st.secrets["gcp_service_account"])
        key  = info.get("private_key", "")
        key  = key.replace("\\n", "\n")
        lines = [l.strip() for l in key.split("\n") if l.strip()]
        if len(lines) >= 3:
            header  = lines[0]
            footer  = lines[-1]
            b64     = "".join(lines[1:-1])
            b64_fmt = "\n".join(b64[i:i+64] for i in range(0, len(b64), 64))
            key     = f"{header}\n{b64_fmt}\n{footer}\n"
        info["private_key"] = key
        creds  = Credentials.from_service_account_info(info, scopes=scopes)
        client = gspread.authorize(creds)
        return client, creds, None
    except Exception as e:
        return None, None, str(e)


# ─────────────────────────────────────────────────────────────────
#  GOOGLE SHEETS
# ─────────────────────────────────────────────────────────────────

@st.cache_data(ttl=300, show_spinner=False)
def load_sheet() -> tuple[list, str | None]:
    client, _creds, err = _get_credentials()
    if err:
        return [], err
    try:
        ws   = client.open_by_key(SHEET_ID).worksheet(FORM_SHEET)
        data = ws.get_all_values()
        return data, None
    except Exception as e:
        return [], str(e)


# ─────────────────────────────────────────────────────────────────
#  GOOGLE DRIVE — ดาวน์โหลดรูปใบเสร็จ
# ─────────────────────────────────────────────────────────────────

def download_drive_image(file_id: str, creds: Credentials) -> bytes | None:
    try:
        if not creds.valid or creds.expired:
            creds.refresh(GoogleAuthRequest())
        url  = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
        resp = requests.get(
            url,
            headers={"Authorization": f"Bearer {creds.token}"},
            timeout=30,
        )
        return resp.content if resp.status_code == 200 else None
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────
#  CLAUDE VISION — อ่านใบเสร็จ
# ─────────────────────────────────────────────────────────────────

def analyze_receipt(image_bytes: bytes) -> dict:
    """Use Claude Vision to extract items/qty/price from receipt image."""
    try:
        api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return {"items": [], "branch": None, "error": "ไม่พบ ANTHROPIC_API_KEY"}

        client = anthropic.Anthropic(api_key=api_key)

        # Detect image type
        if image_bytes[:4] == b"\x89PNG":
            media_type = "image/png"
        elif image_bytes[:2] == b"\xff\xd8":
            media_type = "image/jpeg"
        elif image_bytes[:4] in (b"GIF8", b"GIF9"):
            media_type = "image/gif"
        else:
            media_type = "image/jpeg"

        img_b64 = base64.standard_b64encode(image_bytes).decode()

        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": img_b64,
                            },
                        },
                        {
                            "type": "text",
                            "text": (
                                "อ่านใบเสร็จนี้แล้วตอบ JSON เท่านั้น ไม่ต้องอธิบายเพิ่ม:\n"
                                '{"items": [["ชื่อสินค้า", จำนวน, ราคาต่อหน่วย]], "branch": null}\n\n'
                                "กฎการตั้งชื่อสินค้า: ใช้รูปแบบ {ย่อ}{ปี2หลัก} {ประเภท}\n"
                                "ตัวอย่าง: SW18 ลิ้นหน้า, CV20 สเกิร์ตรอบคัน, YA17 สเกิร์ตข้าง\n"
                                "ย่อ: JAZZ=JZ, VIOS=VO, CIVIC=CV, ATTRAGE=AT, SWIFT=SW, CIAZ=CZ, HR-V=HR, YARIS=YA\n\n"
                                "branch: ถ้าเอกสาร ABBHCT หรือ ABBHPW ให้อ่านชื่อพนักงาน "
                                "เช่น T5→5, T12→12, มิฉะนั้นใส่ null\n"
                                "ราคาต่อหน่วย: ตัวเลขเท่านั้น ไม่มีหน่วย"
                            ),
                        },
                    ],
                }
            ],
        )

        text = response.content[0].text.strip()
        m    = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            result = json.loads(m.group())
            return result
        return {"items": [], "branch": None}
    except Exception as e:
        return {"items": [], "branch": None, "error": str(e)}


# ─────────────────────────────────────────────────────────────────
#  อ่านรูปใบเสร็จทั้งสัปดาห์
# ─────────────────────────────────────────────────────────────────

def read_receipts_for_week(
    all_data: list,
    monday: date,
    creds: Credentials,
) -> dict:
    """Download & analyze all receipt images for the week. Returns updated cache."""
    sunday    = monday + timedelta(days=6)
    rows      = all_data[1:]
    week_rows = [
        r for r in rows
        if (d := parse_date_val(r[1] if len(r) > 1 else None))
        and monday <= d <= sunday
    ]

    # Collect unique file IDs
    file_ids: list[str] = []
    seen: set[str] = set()
    for r in week_rows:
        fid = get_drive_file_id(safe(r, 6))
        if fid and fid not in seen:
            file_ids.append(fid)
            seen.add(fid)

    cache   = dict(st.session_state.get("receipt_cache", {}))
    new_ids = [fid for fid in file_ids if fid not in cache]

    if not new_ids:
        return cache

    prog = st.progress(0.0, text=f"🖼️ กำลังอ่านใบเสร็จ 0/{len(new_ids)}...")
    for i, fid in enumerate(new_ids):
        prog.progress(i / len(new_ids), text=f"🖼️ กำลังอ่านใบเสร็จ {i + 1}/{len(new_ids)}...")
        img = download_drive_image(fid, creds)
        if img:
            result = analyze_receipt(img)
            cache[fid] = {
                "items":  result.get("items", []),
                "branch": result.get("branch"),
            }
        else:
            cache[fid] = {"items": [], "branch": None}

    prog.progress(1.0, text=f"✅ อ่านใบเสร็จครบ {len(new_ids)} ใบ")
    return cache

# ─────────────────────────────────────────────────────────────────
#  BUILD REPORT DataFrame
# ─────────────────────────────────────────────────────────────────

def build_report(
    all_data: list,
    monday: date,
    receipt_cache: dict,
) -> tuple[pd.DataFrame, int]:
    sunday    = monday + timedelta(days=6)
    rows      = all_data[1:]
    week_rows = [
        r for r in rows
        if (d := parse_date_val(r[1] if len(r) > 1 else None))
        and monday <= d <= sunday
    ]

    by_date: dict[date, list] = {}
    for r in week_rows:
        d = parse_date_val(r[1])
        if d:
            by_date.setdefault(d, []).append(r)

    output = []

    for i in range(7):
        day      = monday + timedelta(days=i)
        is_sun   = day.weekday() == 6
        day_rows = by_date.get(day, [])
        th_date  = to_thai_date(day)
        day_name = f"วัน{DAY_TH[day.weekday()]}"

        _base = dict(
            ว_ด_ป=th_date, วัน=day_name,
            เลขที่="", สาขา="", เลขที่เอกสาร="",
            ชื่อลูกค้า="", ที่อยู่="", TAX_ID="",
            รายการสินค้า="", จำนวน=None, ราคา_หน่วย=None,
            รวม=None, VAT_7=None, รวมทั้งสิ้น=None,
            ช่องทาง="", อีเมล_จัดส่ง="",
            _type="",
        )

        if is_sun:
            output.append({**_base, "ชื่อลูกค้า": "— วันอาทิตย์ หยุด —", "_type": "sun"})
            continue

        if not day_rows:
            output.append({**_base, "ชื่อลูกค้า": "(ไม่มีข้อมูล)", "_type": "empty"})
            continue

        # Group by doc_no (1 ใบเสร็จ = 1 doc_no = 1 เลขที่)
        by_doc: dict[str, list] = {}
        order:  list[str]       = []
        for r in day_rows:
            key = safe(r, 2) or f"__no_{len(order)}"
            if key not in by_doc:
                by_doc[key] = []
                order.append(key)
            by_doc[key].append(r)

        for slot, doc_no in enumerate(order):
            r         = by_doc[doc_no][0]
            tax_no    = calc_tax_no(day, slot)
            drive_url = safe(r, 6)
            file_id   = get_drive_file_id(drive_url)

            # Receipt cache lookup
            cached         = receipt_cache.get(file_id, {}) if file_id else {}
            items          = cached.get("items", [])
            branch_override = cached.get("branch")
            branch         = branch_override if branch_override else get_branch(doc_no)

            if items:
                # ─── หลาย rows ต่อ 1 ใบเสร็จ ───────────────────────────
                for item_idx, item in enumerate(items):
                    item_name  = item[0] if len(item) > 0 else ""
                    item_qty   = item[1] if len(item) > 1 else None
                    item_price = item[2] if len(item) > 2 else None

                    # Col A, C-G, N, O เฉพาะ item แรก; เลขที่ทุก row
                    first = item_idx == 0
                    output.append({
                        "ว_ด_ป":          th_date if first else "",
                        "วัน":            day_name if first else "",
                        "เลขที่":         tax_no,
                        "สาขา":           branch if first else "",
                        "เลขที่เอกสาร":   safe(r, 2) if first else "",
                        "ชื่อลูกค้า":     safe(r, 3) if first else "",
                        "ที่อยู่":        safe(r, 4) if first else "",
                        "TAX_ID":         safe(r, 5) if first else "",
                        "รายการสินค้า":   item_name,
                        "จำนวน":          item_qty,
                        "ราคา_หน่วย":     item_price,
                        "รวม": None, "VAT_7": None, "รวมทั้งสิ้น": None,
                        "ช่องทาง":        safe(r, 7) if first else "",
                        "อีเมล_จัดส่ง":   safe(r, 8) if first else "",
                        "_type":          "data",
                    })
            else:
                # ─── 1 row ต่อ 1 ใบเสร็จ (ยังไม่มี cache) ──────────────
                output.append({
                    "ว_ด_ป":         th_date, "วัน": day_name,
                    "เลขที่":        tax_no,  "สาขา": branch,
                    "เลขที่เอกสาร":  safe(r, 2),
                    "ชื่อลูกค้า":    safe(r, 3),
                    "ที่อยู่":       safe(r, 4), "TAX_ID": safe(r, 5),
                    "รายการสินค้า":  "", "จำนวน": None, "ราคา_หน่วย": None,
                    "รวม": None, "VAT_7": None, "รวมทั้งสิ้น": None,
                    "ช่องทาง":       safe(r, 7), "อีเมล_จัดส่ง": safe(r, 8),
                    "_type":         "data",
                })

    rename = {
        "ว_ด_ป": "ว.ด.ป.", "TAX_ID": "TAX ID",
        "ราคา_หน่วย": "ราคา/หน่วย", "VAT_7": "VAT 7%",
        "อีเมล_จัดส่ง": "อีเมล/จัดส่ง",
    }
    df = pd.DataFrame(output).rename(columns=rename)
    return df, len(week_rows)

# ─────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────

def generate_excel(df: pd.DataFrame, monday: date, week_num: int) -> io.BytesIO:
    sunday = monday + timedelta(days=6)
    wb = Workbook()
    ws = wb.active
    ws.title = f"W{week_num:02d}"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(color="000000", bold=False, size=10):
        return Font(name="Sarabun", color=color, bold=bold, size=size)

    thin   = Side(style="thin", color="C0CFE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal="center", vertical="center")
    R = Alignment(horizontal="right",  vertical="center")
    L = Alignment(horizontal="left",   vertical="center")

    NCOLS     = 15
    col_names = [
        "ว.ด.ป.", "เลขที่", "สาขา", "เลขที่เอกสาร", "ชื่อลูกค้า",
        "ที่อยู่", "TAX ID", "รายการสินค้า", "จำนวน", "ราคา/หน่วย",
        "รวม", "VAT 7%", "รวมทั้งสิ้น", "ช่องทาง", "อีเมล/จัดส่ง",
    ]
    col_widths = [12, 11, 6, 16, 18, 22, 13, 22, 7, 10, 10, 9, 11, 14, 22]

    r = 1

    # Row 1 – Title
    title = (
        f"รายงานภาษีขายประจำสัปดาห์ที่ {week_num}  —  "
        f"{to_thai_date(monday)}  ถึง  {to_thai_date(sunday)}"
    )
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    c = ws.cell(1, 1, title)
    c.fill = fill("1A3A6E"); c.font = font("FFFFFF", True, 12); c.alignment = C
    ws.row_dimensions[1].height = 30
    r = 2

    # Row 2 – Headers
    for col, h in enumerate(col_names, 1):
        c = ws.cell(r, col, h)
        c.fill = fill("4472C4"); c.font = font("FFFFFF", True)
        c.alignment = C; c.border = border
    ws.row_dimensions[r].height = 22
    r += 1

    # Data rows
    for _, row_data in df.iterrows():
        row_type = row_data.get("_type", "data")
        qty      = row_data.get("จำนวน")
        price    = row_data.get("ราคา/หน่วย")
        has_num  = pd.notna(qty) and pd.notna(price) and qty and price

        # Excel col order: A=ว.ด.ป., B=เลขที่, C=สาขา, D=เลขที่เอกสาร, ...
        vals = [
            row_data.get("ว.ด.ป.", ""),
            row_data.get("เลขที่", ""),
            row_data.get("สาขา", ""),
            row_data.get("เลขที่เอกสาร", ""),
            row_data.get("ชื่อลูกค้า", ""),
            row_data.get("ที่อยู่", ""),
            row_data.get("TAX ID", ""),
            row_data.get("รายการสินค้า", ""),
            qty   if has_num else "",
            price if has_num else "",
            qty * price                                    if has_num else "",
            round(qty * price * 0.07, 2)                   if has_num else "",
            qty * price + round(qty * price * 0.07, 2)     if has_num else "",
            row_data.get("ช่องทาง", ""),
            row_data.get("อีเมล/จัดส่ง", ""),
        ]

        bg = {"sun": "FFE0E0", "empty": "F8F8F8"}.get(row_type, "FFFFFF")
        fg = "BBBBBB" if row_type == "empty" else "333333"

        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            c.fill      = fill(bg)
            c.font      = font(fg)
            c.border    = border
            c.alignment = (
                C if col in (1, 2, 3) else
                R if col in (9, 10, 11, 12, 13) else
                L
            )
            if col in (11, 12, 13):
                c.number_format = "#,##0.00"

        ws.row_dimensions[r].height = 20
        r += 1

    # Total row
    data_rows = df[df["_type"] == "data"]
    total_k   = sum(
        (q or 0) * (p or 0)
        for q, p in zip(data_rows["จำนวน"], data_rows["ราคา/หน่วย"])
        if pd.notna(q) and pd.notna(p)
    )
    total_l = round(total_k * 0.07, 2)
    total_m = total_k + total_l

    ws.merge_cells(f"A{r}:{get_column_letter(10)}{r}")
    c = ws.cell(r, 1, "รวมทั้งสัปดาห์")
    c.fill = fill("DCE6F7"); c.font = font(bold=True); c.alignment = R

    for col, val in ((11, total_k), (12, total_l), (13, total_m)):
        c = ws.cell(r, col, val or "")
        c.fill = fill("DCE6F7"); c.font = font(bold=True)
        c.alignment = R; c.number_format = "#,##0.00"; c.border = border

    for col in (14, 15):
        ws.cell(r, col).fill = fill("DCE6F7")

    ws.row_dimensions[r].height = 24

    # Column widths & freeze
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────────
#  MAIN UI
# ─────────────────────────────────────────────────────────────────

def main():
    st.title("📊 Weekly Tax Report")
    st.caption("TRC Motorsport — รายงานภาษีขายรายสัปดาห์")
    st.divider()

    # ── Controls ────────────────────────────────────────────────
    col1, col2, col3 = st.columns([3, 1.8, 1.2])

    with col1:
        selected_date = st.date_input(
            "เลือกวันใดก็ได้ในสัปดาห์ที่ต้องการ",
            value=date.today(),
            format="DD/MM/YYYY",
        )

    with col2:
        load_btn = st.button(
            "🔍 โหลดข้อมูล + อ่านรูป",
            type="primary",
            use_container_width=True,
        )

    with col3:
        if st.button("🔄 ล้าง Cache", use_container_width=True):
            load_sheet.clear()
            st.session_state.pop("report_df", None)
            st.session_state.pop("receipt_cache", None)
            st.success("ล้าง cache แล้วค่ะ")
            st.rerun()

    # Calculate week
    iso    = selected_date.isocalendar()
    monday = date.fromisocalendar(iso[0], iso[1], 1)

    # ── Load ────────────────────────────────────────────────────
    if load_btn:
        # 1. Load Google Sheet
        with st.spinner("📋 ดึงข้อมูลจาก Google Sheet..."):
            all_data, err = load_sheet()

        if err:
            st.error(f"❌ {err}")
            st.info(
                "💡 ไปที่ **Settings → Secrets** ใน Streamlit Cloud แล้วเพิ่ม `[gcp_service_account]`\n"
                "ดูตัวอย่างใน `secrets.toml.example` ค่ะ"
            )
            return

        # 2. อ่านรูปใบเสร็จจาก Drive ด้วย Claude Vision
        _client, creds, _err = _get_credentials()
        has_api_key = bool(st.secrets.get("ANTHROPIC_API_KEY", ""))

        if creds and has_api_key:
            receipt_cache = read_receipts_for_week(all_data, monday, creds)
            st.session_state["receipt_cache"] = receipt_cache
        else:
            receipt_cache = st.session_state.get("receipt_cache", {})
            if not has_api_key:
                st.warning(
                    "⚠️ ไม่พบ **ANTHROPIC_API_KEY** ใน Secrets — "
                    "รายการสินค้า/จำนวน/ราคา จะต้องกรอกเองค่ะ"
                )

        # 3. Build report DataFrame
        report_df, n = build_report(all_data, monday, receipt_cache)

        st.session_state["report_df"]  = report_df
        st.session_state["monday"]     = monday
        st.session_state["week_num"]   = iso[1]
        st.session_state["raw_count"]  = n
        st.session_state["all_data"]   = all_data

    # ── Display ─────────────────────────────────────────────────
    if "report_df" not in st.session_state:
        st.info("👆 เลือกสัปดาห์แล้วกด **โหลดข้อมูล + อ่านรูป** ค่ะ")
        return

    df       = st.session_state["report_df"]
    monday   = st.session_state["monday"]
    week_num = st.session_state["week_num"]
    sunday   = monday + timedelta(days=6)

    # Stats
    m1, m2, m3 = st.columns(3)
    m1.metric("สัปดาห์ที่", week_num)
    m2.metric("ช่วงวันที่", f"{to_thai_date(monday)} – {to_thai_date(sunday)}")
    m3.metric("รายการที่พบ", st.session_state["raw_count"])

    # ── Editable Table ──────────────────────────────────────────
    disabled_cols = [
        "ว.ด.ป.", "วัน", "เลขที่", "สาขา", "เลขที่เอกสาร",
        "ชื่อลูกค้า", "ที่อยู่", "TAX ID",
        "รวม", "VAT 7%", "รวมทั้งสิ้น",
        "ช่องทาง", "อีเมล/จัดส่ง",
    ]

    col_config = {
        "ว.ด.ป.":          st.column_config.TextColumn("ว.ด.ป. 🔒",          disabled=True),
        "วัน":             st.column_config.TextColumn("วัน 🔒",              disabled=True, width="small"),
        "เลขที่":          st.column_config.TextColumn("เลขที่ 🔒",           disabled=True, width="small"),
        "สาขา":            st.column_config.TextColumn("สาขา 🔒",             disabled=True, width="small"),
        "เลขที่เอกสาร":    st.column_config.TextColumn("เลขที่เอกสาร 🔒",    disabled=True),
        "ชื่อลูกค้า":      st.column_config.TextColumn("ชื่อลูกค้า 🔒",      disabled=True),
        "ที่อยู่":         st.column_config.TextColumn("ที่อยู่ 🔒",          disabled=True),
        "TAX ID":          st.column_config.TextColumn("TAX ID 🔒",           disabled=True),
        "รายการสินค้า":    st.column_config.TextColumn("รายการสินค้า",        width="medium"),
        "จำนวน":           st.column_config.NumberColumn("จำนวน",             min_value=0, step=1,   format="%d"),
        "ราคา/หน่วย":      st.column_config.NumberColumn("ราคา/หน่วย",        min_value=0,           format="%g"),
        "รวม":             st.column_config.NumberColumn("รวม 🔒",            disabled=True,         format="%.2f"),
        "VAT 7%":          st.column_config.NumberColumn("VAT 7% 🔒",         disabled=True,         format="%.2f"),
        "รวมทั้งสิ้น":     st.column_config.NumberColumn("รวมทั้งสิ้น 🔒",   disabled=True,         format="%.2f"),
        "ช่องทาง":         st.column_config.TextColumn("ช่องทาง 🔒",          disabled=True),
        "อีเมล/จัดส่ง":    st.column_config.TextColumn("อีเมล/จัดส่ง 🔒",   disabled=True),
    }

    edited = st.data_editor(
        df[DISPLAY_COLS],
        column_config=col_config,
        disabled=disabled_cols,
        hide_index=True,
        use_container_width=True,
        key="tax_editor",
        num_rows="fixed",
    )

    # Recalculate totals from edited จำนวน / ราคา/หน่วย
    edited["รวม"] = edited.apply(
        lambda x: (x["จำนวน"] or 0) * (x["ราคา/หน่วย"] or 0)
        if pd.notna(x["จำนวน"]) and pd.notna(x["ราคา/หน่วย"])
        and x["จำนวน"] and x["ราคา/หน่วย"]
        else None,
        axis=1,
    )
    edited["VAT 7%"]      = edited["รวม"].apply(lambda x: round(x * 0.07, 2) if x else None)
    edited["รวมทั้งสิ้น"] = edited.apply(
        lambda x: (x["รวม"] or 0) + (x["VAT 7%"] or 0) if x["รวม"] else None,
        axis=1,
    )

    # Summary metrics
    total_k = edited["รวม"].sum() or 0
    if total_k > 0:
        st.divider()
        t1, t2, t3 = st.columns(3)
        t1.metric("💰 รวมยอดขาย",  f"฿{total_k:,.2f}")
        t2.metric("🧾 VAT 7%",      f"฿{edited['VAT 7%'].sum():,.2f}")
        t3.metric("✅ รวมทั้งสิ้น", f"฿{edited['รวมทั้งสิ้น'].sum():,.2f}")

    # ── Export ──────────────────────────────────────────────────
    st.divider()
    export_df          = edited.copy()
    export_df["_type"] = df["_type"].values

    excel_buf = generate_excel(export_df, monday, week_num)

    st.download_button(
        label="⬇️ Export Excel (.xlsx)",
        data=excel_buf,
        file_name=f"TaxReport_W{week_num:02d}_{monday.year}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    main()
